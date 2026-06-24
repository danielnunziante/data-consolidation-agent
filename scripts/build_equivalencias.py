"""Genera `app/utils/seccion_equivalencias.py` desde `Ejemplos Data/EQUIVALENCIAS.xlsx`.

El Excel es la *fuente de verdad* que mantiene el cliente. Este script lo
convierte a un módulo Python versionable (sin dependencias de archivos en
runtime ni cambios en el .spec de PyInstaller) que consume el normalizador
`app/utils/seccion.py`.

Uso:
    python scripts/build_equivalencias.py
    python scripts/build_equivalencias.py --xlsx "ruta/al/EQUIVALENCIAS.xlsx"

Columnas esperadas en la hoja `DATOS`: CODIGO | COMPAÑÍA | SECCION
"""
from __future__ import annotations

import argparse
from datetime import date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "Ejemplos Data" / "EQUIVALENCIAS.xlsx"
OUT_PY = ROOT / "app" / "utils" / "seccion_equivalencias.py"


def _read_rows(xlsx: Path) -> list[tuple[str, str, str]]:
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    ws = wb["DATOS"] if "DATOS" in wb.sheetnames else wb.worksheets[0]

    rows: list[tuple[str, str, str]] = []
    for codigo, compania, seccion, *_ in ws.iter_rows(min_row=2, values_only=True):
        if compania is None or seccion is None:
            continue
        # Los códigos numéricos llegan como int/float; los guardamos como texto
        # para que el módulo generado sea estable y legible.
        if isinstance(codigo, float) and codigo.is_integer():
            codigo = int(codigo)
        cod = "" if codigo is None else str(codigo).strip()
        rows.append((cod, str(compania).strip(), str(seccion).strip()))
    return rows


def _render(rows: list[tuple[str, str, str]], src_name: str) -> str:
    lines = [
        '"""Equivalencias de SECCION por compañía (GENERADO AUTOMÁTICAMENTE).',
        "",
        f"NO EDITAR A MANO. Generado por scripts/build_equivalencias.py desde",
        f"'{src_name}' el {date.today().isoformat()}.",
        "",
        "Cada tupla es (CODIGO_ORIGEN, COMPAÑÍA, SECCION_NORMALIZADA). El CODIGO",
        "es el valor crudo tal como lo emite el parser de esa compañía.",
        '"""',
        "from __future__ import annotations",
        "",
        "EQUIVALENCIAS_RAW: list[tuple[str, str, str]] = [",
    ]
    for cod, comp, sec in rows:
        lines.append(f"    ({cod!r}, {comp!r}, {sec!r}),")
    lines.append("]")
    lines.append("")
    return "\n".join(lines)


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Ruta al EQUIVALENCIAS.xlsx")
    ap.add_argument("--out", default=str(OUT_PY), help="Ruta del módulo .py a generar")
    args = ap.parse_args()

    xlsx = Path(args.xlsx)
    if not xlsx.exists():
        raise SystemExit(f"No se encontró el Excel: {xlsx}")

    rows = _read_rows(xlsx)
    out = Path(args.out)
    out.write_text(_render(rows, xlsx.name), encoding="utf-8")
    print(f"OK: {len(rows)} equivalencias -> {out}")


if __name__ == "__main__":
    main()
