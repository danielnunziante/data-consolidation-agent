from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from app.config import BASE_COLUMNS, BASE_SHEET_NAME
from app.models import Record
from app.workbook_builder import build_workbook, sort_records


def _rec(**overrides) -> Record:
    base = dict(
        fecha=date(2026, 3, 1),
        poliza="1",
        asegurado="X",
        seccion="S",
        compania="ALLIANZ",
        tipo="PR",
        comisiones=10.0,
        prima=100.0,
        premio=110.0,
        source_file="x.xlsx",
    )
    base.update(overrides)
    return Record(**base)


def test_sort_records_respects_company_order():
    records = [
        _rec(compania="ZURICH", poliza="9"),
        _rec(compania="ALLIANZ", poliza="3"),
        _rec(compania="ALLIANZ", poliza="1"),
    ]
    s = sort_records(records)
    assert [r.compania for r in s] == ["ALLIANZ", "ALLIANZ", "ZURICH"]
    assert [r.poliza for r in s[:2]] == ["1", "3"]


def test_build_workbook_generates_base_sheet(tmp_path: Path):
    records = [_rec(poliza="12345"), _rec(poliza="12345"), _rec(poliza="99999", compania="ZURICH")]
    out = tmp_path / "out.xlsx"
    build_workbook(records, str(out), date(2026, 3, 1))

    assert out.exists()
    wb = load_workbook(out)
    assert wb.sheetnames == [BASE_SHEET_NAME]
    ws = wb[BASE_SHEET_NAME]
    headers = [c.value for c in ws[1]]
    assert headers == BASE_COLUMNS
    assert ws.max_row == len(records) + 1
    assert ws.freeze_panes == "A2"
    # autofilter
    assert ws.auto_filter.ref is not None
    # Fórmula de duplicados en Columna1
    formula_cell = ws.cell(row=2, column=len(BASE_COLUMNS))
    assert formula_cell.value and str(formula_cell.value).startswith("=IF(COUNTIF(")
