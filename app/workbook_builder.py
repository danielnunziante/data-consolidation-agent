"""Genera el workbook `BASE` final desde cero."""
from __future__ import annotations

from datetime import date
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import BASE_COLUMNS, BASE_SHEET_NAME, COMPANY_ORDER
from .models import Record
from .utils.excel_styles import (
    BODY_ALIGN,
    BODY_FONT,
    COLUMN_WIDTHS,
    DATE_FMT,
    HEADER_ALIGN,
    HEADER_BORDER,
    HEADER_FILL,
    HEADER_FONT,
    MONEY_FMT,
)


def _company_rank(compania: str) -> int:
    try:
        return COMPANY_ORDER.index(compania)
    except ValueError:
        return len(COMPANY_ORDER) + 1


def sort_records(records: list[Record]) -> list[Record]:
    def key(r: Record):
        return (
            _company_rank(r.compania),
            r.compania,
            (r.poliza or ""),
            r.tipo or "",
            (r.asegurado or ""),
        )

    return sorted(records, key=key)


def _write_header(ws: Worksheet) -> None:
    for col_idx, name in enumerate(BASE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER

    # anchos
    for col_idx, name in enumerate(BASE_COLUMNS, start=1):
        width = COLUMN_WIDTHS.get(name, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28


def _write_body(ws: Worksheet, records: list[Record], fecha: date) -> None:
    # Índices 1-based de columnas
    col_idx = {name: i + 1 for i, name in enumerate(BASE_COLUMNS)}

    for r_i, rec in enumerate(records, start=2):
        ws.cell(row=r_i, column=col_idx["FECHA"], value=fecha).number_format = DATE_FMT
        ws.cell(row=r_i, column=col_idx["POLIZA"], value=rec.poliza)
        ws.cell(row=r_i, column=col_idx["ASEGURADO"], value=rec.asegurado)
        ws.cell(row=r_i, column=col_idx["SECCION"], value=rec.seccion)
        ws.cell(row=r_i, column=col_idx["COMPAÑÍA"], value=rec.compania)
        ws.cell(row=r_i, column=col_idx["TIPO"], value=rec.tipo)

        c_com = ws.cell(row=r_i, column=col_idx["COMISIONES"], value=rec.comisiones)
        c_pri = ws.cell(row=r_i, column=col_idx["PRIMA"], value=rec.prima)
        c_pre = ws.cell(row=r_i, column=col_idx["PREMIO"], value=rec.premio)
        for c in (c_com, c_pri, c_pre):
            c.number_format = MONEY_FMT

        # COMPROBACION DE DUPLICADOS permanece vacío (spec)
        # Columna1: fórmula de control de duplicados dentro de BASE
        poliza_cell = f"B{r_i}"
        poliza_col = col_idx["POLIZA"]
        range_ref = f"${get_column_letter(poliza_col)}$2:${get_column_letter(poliza_col)}${len(records) + 1}"
        formula = f'=IF(COUNTIF({range_ref},{poliza_cell})>1,"VER","")'
        ws.cell(row=r_i, column=col_idx["Columna1"], value=formula)

        for c_name in BASE_COLUMNS:
            cell = ws.cell(row=r_i, column=col_idx[c_name])
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN


def build_workbook(
    records: Iterable[Record],
    output_file: str,
    period_first_day: date,
) -> None:
    """Crea un workbook .xlsx desde cero con la hoja BASE."""
    records = sort_records(list(records))

    wb = Workbook()
    # Borrar hoja default y crear BASE
    default_ws = wb.active
    wb.remove(default_ws)
    ws = wb.create_sheet(BASE_SHEET_NAME)

    _write_header(ws)
    _write_body(ws, records, period_first_day)

    # Congelar fila superior
    ws.freeze_panes = "A2"

    # Autofiltro completo
    last_col = get_column_letter(len(BASE_COLUMNS))
    last_row = max(1, len(records) + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Asegurar carpeta
    out = Path(output_file).expanduser().resolve()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
